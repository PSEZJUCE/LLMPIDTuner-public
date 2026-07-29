from pathlib import Path

import openpyxl

from llmpidtuner.cli import _parse_group_spec
from llmpidtuner.config import CaseConfig
from llmpidtuner.models import PIDParams, SimulationSettings
from llmpidtuner.runner import (
    _batch_folder_name,
    _batch_plants,
    _run_batch,
    write_batch_results_from_config,
)


def test_parse_group_spec_supports_ranges_and_single_groups():
    assert _parse_group_spec("1-3,7,10-11") == {1, 2, 3, 7, 10, 11}


def test_collect_batch_rebuilds_workbook_from_group_status(tmp_path: Path):
    plants_path = tmp_path / "plants.yaml"
    plants_path.write_text(
        "plants:\n- group: 1\n  k: 0.5\n  t: 10\n- group: 2\n  k: 0.6\n  t: 20\n",
        encoding="utf-8",
    )
    config = CaseConfig(
        name="parallel_eval",
        system="first_order_batch",
        mode="llm",
        output_dir=str(tmp_path / "runs"),
        initial_pid=PIDParams(1.0, 0.1, 0.01),
        simulation=SimulationSettings(sim_time=10.0, num_points=101, time_delay=0.0),
        batch={"plants_path": str(plants_path)},
    )
    group, plant = _batch_plants(config.batch or {}, "first_order")[0]
    status_dir = Path(config.output_dir) / config.name / _batch_folder_name(group, plant)
    status_dir.mkdir(parents=True)
    (status_dir / "run_status.yaml").write_text(
        "status: converged\n"
        "stop_reason: success_threshold\n"
        "completed_iterations: 2\n"
        "failed_next_llm_call: null\n"
        "final_pid:\n  kp: 1.2\n  ki: 0.2\n  kd: 0.03\n"
        "final_iae: 4.0\n"
        "final_overshoot: 3.0\n"
        "final_steady_state_error: 0.1\n",
        encoding="utf-8",
    )

    workbook_path = write_batch_results_from_config(config)

    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    status_column = headers.index("Run Status") + 1
    final_kp_column = headers.index("Final Kp") + 1
    assert worksheet.cell(row=2, column=status_column).value == "converged"
    assert worksheet.cell(row=2, column=final_kp_column).value == 1.2
    assert worksheet.cell(row=3, column=status_column).value == "missing"

def test_run_batch_limits_work_to_selected_groups_without_workbook(tmp_path: Path):
    plants_path = tmp_path / "plants.yaml"
    plants_path.write_text(
        "plants:\n- group: 1\n  k: 0.5\n  t: 10\n- group: 2\n  k: 0.6\n  t: 20\n",
        encoding="utf-8",
    )
    config = CaseConfig(
        name="sharded_eval",
        system="first_order_batch",
        mode="dry_run",
        output_dir=str(tmp_path / "runs"),
        initial_pid=PIDParams(1.0, 0.1, 0.01),
        simulation=SimulationSettings(sim_time=10.0, num_points=101, time_delay=0.0),
        batch={"plants_path": str(plants_path)},
    )
    plants = _batch_plants(config.batch or {}, "first_order")

    _run_batch(
        config,
        Path(config.output_dir) / config.name,
        "first_order",
        batch_groups={2},
        write_batch_excel=False,
    )

    first_folder = Path(config.output_dir) / config.name / _batch_folder_name(*plants[0])
    second_folder = Path(config.output_dir) / config.name / _batch_folder_name(*plants[1])
    assert not first_folder.exists()
    assert (second_folder / "pid_tuning_comparison.png").exists()
    assert not (Path(config.output_dir) / config.name / "experiment_results.xlsx").exists()